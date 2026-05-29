"""Build and load the classification dictionary from the master xlsx."""
import json
import os
import re
import statistics
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import openpyxl

from .normalize import make_key, normalize_key


_TRAILING_SEP = re.compile(r"[\s\-–—]+$")


def _common_base(values) -> str:
    """Longest common prefix of obs variants, trimmed of trailing separators.

    When a recurring value maps to several observação variants that share a stem
    (e.g. 'Contabilidade - PS - {N}ª Emissão'), we keep the stem and leave the
    distinguishing suffix for the contadora to complete.
    """
    cleaned = [v.strip() for v in values if isinstance(v, str) and v.strip()]
    if not cleaned:
        return ""
    base = os.path.commonprefix(cleaned)
    return _TRAILING_SEP.sub("", base).strip()


DETAIL_SHEETS = ["AECO", "SEC", "C6", "TECH", "Conta Simples"]
CLASS_COLS = ("descricao", "observacoes", "fluxo_caixa", "empresa")


def _classification_tuple(descr, obs, fluxo, empresa) -> tuple:
    return (
        (descr or "").strip() if isinstance(descr, str) else descr,
        (obs or "").strip() if isinstance(obs, str) else obs,
        (fluxo or "").strip() if isinstance(fluxo, str) else fluxo,
        (empresa or "").strip() if isinstance(empresa, str) else empresa,
    )


def _classification_dict(t: tuple) -> dict:
    return dict(zip(CLASS_COLS, t))


def _iter_master_rows(wb):
    """Yield (key, valor, classification_tuple) from every detail sheet."""
    for sn in DETAIL_SHEETS:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        header_row = None
        for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
            if r and r[0] == "Data":
                header_row = i
                break
        if not header_row:
            continue
        for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not r or not r[0]:
                continue
            tipo = r[1]
            benef = r[2]
            descr = r[3]
            obs = r[4]
            fluxo = r[5]
            valor = r[6]
            empresa = r[7]
            if not tipo or empresa is None:
                continue
            key = make_key(str(tipo), str(benef or ""))
            cls = _classification_tuple(descr, obs, fluxo, empresa)
            try:
                v = float(valor) if valor is not None else 0.0
            except (TypeError, ValueError):
                v = 0.0
            yield key, v, cls


def _decide_mode(values_by_class: dict[tuple, list[float]]) -> dict:
    """Decide mode for a key given its observed (classification -> list of values).

    Algorithm:
        1. EXACT: dominant single classification (>=85% of occurrences).
        2. VALUE_BRACKETS: group occurrences by rounded value. For each value group,
           if there's a dominant classification (>=80%), it becomes a bucket.
           Need >=2 buckets and >=70% of all occurrences covered to qualify.
        3. AMBIGUOUS: otherwise — top-5 alternatives.
    """
    total = sum(len(v) for v in values_by_class.values())
    classes = sorted(values_by_class.items(), key=lambda kv: -len(kv[1]))
    top_cls, top_vals = classes[0]
    consistency = len(top_vals) / total

    if len(classes) == 1 or consistency >= 0.85:
        return {
            "mode": "exact",
            "n": total,
            "consistency": round(consistency, 3),
            "classification": _classification_dict(top_cls),
        }

    # Try value-brackets: bucket by rounded value, compute mode per field independently.
    # A bucket is valid if at least 3 of 4 fields have a dominant value (>=70%).
    # We track which fields are consistent vs which the contadora must fill.
    by_value_field: dict[float, dict[str, Counter]] = defaultdict(
        lambda: {c: Counter() for c in CLASS_COLS}
    )
    for cls, vals in values_by_class.items():
        descr, obs, fluxo, empresa = cls
        for v in vals:
            vk = round(v, 2)
            by_value_field[vk]["descricao"][descr] += 1
            by_value_field[vk]["observacoes"][obs] += 1
            by_value_field[vk]["fluxo_caixa"][fluxo] += 1
            by_value_field[vk]["empresa"][empresa] += 1

    buckets = []
    covered = 0
    for val, field_counts in sorted(
        by_value_field.items(), key=lambda kv: -sum(kv[1]["empresa"].values())
    ):
        n_at_val = sum(field_counts["empresa"].values())
        if n_at_val < 2:
            continue
        bucket_cls = {}
        review_fields = []
        observacoes_variants = None
        confident_fields = 0
        for fname in CLASS_COLS:
            counts = field_counts[fname]
            if not counts:
                continue
            top_v, top_n = counts.most_common(1)[0]
            if top_n / n_at_val >= 0.70:
                bucket_cls[fname] = top_v
                confident_fields += 1
            elif fname == "observacoes":
                # Value recurs but the observação suffix rotates (e.g. emission
                # number). Fill the shared stem and hand the rest to the contadora.
                review_fields.append(fname)
                bucket_cls[fname] = _common_base(counts.keys())
                observacoes_variants = [
                    {"observacoes": v, "n": c}
                    for v, c in counts.most_common()
                    if v
                ]
            else:
                review_fields.append(fname)
                bucket_cls[fname] = top_v  # best guess; UI marks for review
        if confident_fields >= 3:
            bucket = {
                "valor_aprox": val,
                "tolerance": 0.01,
                "n": n_at_val,
                "confident_fields": confident_fields,
                "classification": bucket_cls,
            }
            if review_fields:
                bucket["review_fields"] = review_fields
            if observacoes_variants:
                bucket["observacoes_variants"] = observacoes_variants
            buckets.append(bucket)
            covered += n_at_val

    if len(buckets) >= 2 and covered / total >= 0.50:
        return {
            "mode": "value_brackets",
            "n": total,
            "buckets": buckets,
            "fallback_to_llm": True,
        }

    # ambiguous: keep top-5 with example valores
    top_alts = []
    for cls, vals in classes[:5]:
        top_alts.append({
            "n": len(vals),
            "classification": _classification_dict(cls),
            "example_valores": [round(v, 2) for v in vals[:3]],
        })
    return {
        "mode": "ambiguous",
        "n": total,
        "top_alternatives": top_alts,
    }


def build(master_path: str | Path) -> dict:
    wb = openpyxl.load_workbook(master_path, data_only=True)

    # key -> classification_tuple -> list of valores
    grouped: dict[str, dict[tuple, list[float]]] = defaultdict(lambda: defaultdict(list))
    fluxos = Counter()
    empresas = Counter()
    descricoes = Counter()
    observacoes = Counter()

    for key, valor, cls in _iter_master_rows(wb):
        descr, obs, fluxo, empresa = cls
        if descr: descricoes[descr] += 1
        if obs: observacoes[obs] += 1
        if fluxo: fluxos[fluxo] += 1
        if empresa: empresas[empresa] += 1
        grouped[key][cls].append(valor)

    entries = {}
    for key, by_cls in grouped.items():
        entries[key] = _decide_mode(by_cls)

    return {
        "version": "1",
        "categories": {
            "fluxo_caixa": sorted(fluxos.keys()),
            "empresa": sorted(empresas.keys()),
        },
        "all_descricoes": sorted(descricoes.keys()),
        "all_observacoes": sorted(observacoes.keys()),
        "frequency": {
            "descricoes": dict(descricoes.most_common()),
            "observacoes": dict(observacoes.most_common()),
        },
        "entries": entries,
    }


def load(path: str | Path) -> dict:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def save(data: dict, path: str | Path) -> None:
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
