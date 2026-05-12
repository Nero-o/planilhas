"""Export classified transactions to the AECO consolidated workbook.

Two output modes:

* ``to_xlsx(txs_df, saldos)`` — legacy: creates a fresh workbook with one tab
  per destination sheet ("AECO", "SEC", "TECH", "Conta Simples", "C6") and a
  "Validação" tab. Header lives on row 1, matching the master.
* ``to_xlsx_overlay(txs_df, master_path, saldos)`` — opens the master
  workbook in place, finds the last data row of each destination sheet, and
  appends the new classified rows. Existing master rows stay locked; the
  newly-appended rows are unlocked so the contadora can edit them. Sheet
  protection is enabled so the lock is enforced.

Routing:
    - source=conta_simples -> "Conta Simples"
    - source=c6            -> "C6"
    - else, by empresa:
        AECO/PS/Cons/Matriz/Bravo/Igor -> "AECO"
        Sec -> "SEC"
        Tech -> "TECH"
"""
import io
from datetime import datetime, timedelta

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Protection


CONF_FILL = {
    "green": PatternFill("solid", start_color="D4F7D4", end_color="D4F7D4"),
    "yellow": PatternFill("solid", start_color="FFF3A3", end_color="FFF3A3"),
    "red": PatternFill("solid", start_color="FFBBBB", end_color="FFBBBB"),
}

EMPRESA_TO_SHEET = {
    "AECO": "AECO",
    "PS": "AECO",
    "Cons": "AECO",
    "Matriz": "AECO",
    "Bravo": "AECO",
    "Igor": "AECO",
    "Sec": "SEC",
    "Tech": "TECH",
}

SHEETS = ["AECO", "SEC", "TECH", "Conta Simples", "C6"]
HEADER = [
    "Data", "Tipo de Pagamento", "Beneficiário", "Descrição",
    "Observações", "Fluxo de Caixa", "Valor", "Empresa",
]
HEADER_C6 = [
    "Data", "Tipo de Pagamento", "Beneficiário", "Descrição",
    "Observações", "Fluxo de Caixa", "Entrada", "Empresa",
]


def route_row(row) -> str:
    if row["source"] == "conta_simples":
        return "Conta Simples"
    if row["source"] == "c6":
        return "C6"
    return EMPRESA_TO_SHEET.get(row["empresa"], "AECO")


def _coerce_date_cell(data_val):
    if isinstance(data_val, datetime):
        return data_val
    if isinstance(data_val, (int, float)):
        return datetime(1899, 12, 30) + timedelta(days=float(data_val))
    if isinstance(data_val, str):
        try:
            return datetime.fromisoformat(data_val[:10])
        except ValueError:
            return data_val
    return data_val


def _write_validation_sheet(wb, txs_df, saldos: dict | None):
    ws = wb.create_sheet("Validação", 0)
    ws.append(["Fonte", "Saldo inicial", "Saldo final", "Soma transações", "Diferença", "OK?"])
    for c in ws[1]:
        c.font = Font(bold=True)
    if not saldos:
        return
    for src, s in saldos.items():
        if not s or s.get("saldo_inicial") is None or s.get("saldo_final") is None:
            continue
        source_name = "conta_simples" if src == "cs" else src
        soma = float(txs_df.loc[txs_df.source == source_name, "valor"].sum())
        expected = s["saldo_final"] - s["saldo_inicial"]
        diff = soma - expected
        ws.append([
            src, s["saldo_inicial"], s["saldo_final"],
            round(soma, 2), round(diff, 2), "OK" if abs(diff) <= 0.02 else "FALHOU",
        ])
    ws.append([])
    ws.append(["Distribuição por aba"])
    counts = txs_df.apply(route_row, axis=1).value_counts().to_dict()
    for sheet, n in counts.items():
        ws.append([sheet, n])


def _append_tx_row(ws, r, unlock: bool):
    """Append a single transaction row to ``ws``; mark cells unlocked if requested."""
    data_val = _coerce_date_cell(r.data)
    cells = [
        data_val,
        r.tipo,
        r.beneficiario,
        r.descricao or "",
        r.observacoes or "",
        r.fluxo_caixa or "",
        r.valor,
        r.empresa or "",
    ]
    ws.append(cells)
    row_idx = ws.max_row
    fill = CONF_FILL.get(getattr(r, "confidence", None))
    for c in ws[row_idx]:
        if unlock:
            c.protection = Protection(locked=False)
        if fill is not None:
            c.fill = fill


def to_xlsx(txs_df: pd.DataFrame, saldos: dict | None = None) -> bytes:
    """Legacy: build a fresh workbook with the 5 detail sheets + Validação."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _write_validation_sheet(wb, txs_df, saldos or {})

    routed = txs_df.copy()
    routed["_sheet"] = routed.apply(route_row, axis=1)

    for sheet_name in SHEETS:
        ws = wb.create_sheet(sheet_name)
        header = HEADER_C6 if sheet_name == "C6" else HEADER
        # mirror the master: rows 1-2 left blank, header on row 3
        ws.append([])
        ws.append([])
        ws.append(header)
        for c in ws[3]:
            c.font = Font(bold=True)
        rows = routed[routed._sheet == sheet_name]
        for r in rows.itertuples():
            _append_tx_row(ws, r, unlock=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _find_last_data_row(ws) -> int:
    """Last row that has a non-empty ``Data`` column (header is row 1)."""
    last = 1
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
        if row and row[0].value not in (None, ""):
            last = row[0].row
    return last


def to_xlsx_overlay(
    txs_df: pd.DataFrame, master_path, saldos: dict | None = None,
    protect_password: str | None = None,
) -> bytes:
    """Open the master workbook and append new entries, locking existing rows.

    Existing cells in the master are left untouched (they keep their default
    ``locked=True`` protection). Newly appended cells are marked
    ``locked=False`` and sheet protection is enabled so Excel enforces the
    distinction.

    The "Validação" tab is recreated/overwritten with the current run's saldos.
    """
    wb = openpyxl.load_workbook(master_path)

    # Refresh the Validação tab
    if "Validação" in wb.sheetnames:
        del wb["Validação"]
    _write_validation_sheet(wb, txs_df, saldos or {})

    routed = txs_df.copy()
    routed["_sheet"] = routed.apply(route_row, axis=1)

    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            header = HEADER_C6 if sheet_name == "C6" else HEADER
            ws.append(header)
            for c in ws[1]:
                c.font = Font(bold=True)
        ws = wb[sheet_name]
        rows = routed[routed._sheet == sheet_name]
        if rows.empty:
            continue
        # Append after the last data row (preserves master rows unchanged).
        for r in rows.itertuples():
            _append_tx_row(ws, r, unlock=True)
        ws.protection.sheet = True
        ws.protection.enable()
        if protect_password:
            ws.protection.set_password(protect_password)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
