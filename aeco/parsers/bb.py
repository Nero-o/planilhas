"""Parser for Banco do Brasil extratos.

Two known XLSX layouts are supported:

1. **Extrato Conta** (Jan/Feb/Nov/Dez) — same shape as Sicoob:
   Row 1 header ``Data | Lançamento | Detalhes | N° documento | Valor | Tipo Lançamento``;
   ``Saldo Anterior`` on row 2 and ``Saldo do dia`` rows interleaved; values
   formatted as PT-BR money strings with C/D suffix.

2. **Extrato** (Outubro) — legacy "movimento" layout:
   Row 1 = title, row 2 = agência/conta, row 3 = header
   ``Data | observacao | Data balancete | Agencia Origem | Lote | Numero Documento | Cod. Historico | Historico | Valor R$``.
   Values are plain positive numbers; ``Saldo Anterior`` (row near top) and
   ``S A L D O`` (final row, código 999) bracket the period. Sign is derived
   from the ``Historico`` description.

Both produce rows with ``source = "bb"``, which routes them into the SEC tab
(the master's BB account tab). The ``Empresa`` classification does not affect
routing.
"""
from datetime import datetime
import io
import re

import openpyxl
import pandas as pd

from ..normalize import normalize_text, parse_pt_money
from ._csv_io import RowSheet, is_xlsx, read_csv_rows


SALDO_TIPOS = {"Saldo Anterior", "Saldo do dia", "S A L D O"}

_PIX_RE = re.compile(
    r"^\s*(?:\d{2}/\d{2}\s+)?(?:\d{2}:\d{2}\s+)?(?:\d{14}\s+|\d{11}\s+)?(.+?)\s*$"
)
_TED_RE = re.compile(
    r"^\s*\d{3,4}\s+\d{1,4}\s+(?:\d{14}\s+|\d{11}\s+)?(.+?)\s*$"
)

# "Pix - Enviado", "TED-Crédito em Conta", etc.
_NEG_KEYWORDS = (
    "enviado", "saída", "saida", "pagamento", "tarifa", "débito", "debito",
    "imposto", "darf", "das", "boleto", "compra",
)
_POS_KEYWORDS = ("recebido", "crédito", "credito", "entrada", "aporte")


def _normalize_lanc(lanc: str) -> str:
    return normalize_text(str(lanc or "").replace(" - ", " "))


def _extract_beneficiario(lancamento: str, detalhes: str) -> str:
    if not detalhes:
        return ""
    detalhes = detalhes.strip()
    lanc_lower = (lancamento or "").lower()
    if "pix" in lanc_lower:
        m = _PIX_RE.match(detalhes)
        return normalize_text(m.group(1)) if m else normalize_text(detalhes)
    if "ted" in lanc_lower:
        m = _TED_RE.match(detalhes)
        return normalize_text(m.group(1)) if m else normalize_text(detalhes)
    return normalize_text(detalhes)


def _coerce_date(x) -> datetime:
    if isinstance(x, datetime):
        return x
    if isinstance(x, (int, float)):
        return datetime(1899, 12, 30) + pd.Timedelta(days=float(x))
    return datetime.strptime(str(x).strip(), "%d/%m/%Y")


def _coerce_money(x) -> float:
    if x is None or x == "":
        return 0.0
    if isinstance(x, (int, float)):
        return round(float(x), 2)
    s = str(x).strip()
    try:
        return round(float(s), 2)
    except ValueError:
        return parse_pt_money(s)


def _sign_from_historico(hist: str) -> int:
    h = (hist or "").lower()
    if any(k in h for k in _NEG_KEYWORDS):
        return -1
    if any(k in h for k in _POS_KEYWORDS):
        return 1
    return 1


def _load_ws(path):
    if hasattr(path, "read"):
        if hasattr(path, "seek"):
            path.seek(0)
        wb = openpyxl.load_workbook(io.BytesIO(path.read()), data_only=True)
    else:
        wb = openpyxl.load_workbook(path, data_only=True)
    return wb


def _parse_extrato_conta(ws) -> tuple[pd.DataFrame, dict]:
    """Sicoob-shaped sheet: header on row 1; saldos on Saldo Anterior / S A L D O rows."""
    saldo_ini = saldo_fim = None
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 5:
            continue
        data_str, lanc, detalhes, doc, valor_str = row[:5]
        tipo_lanc = row[5] if len(row) > 5 else None
        if not lanc:
            continue
        lanc_str = str(lanc).strip()
        if lanc_str == "Saldo Anterior":
            saldo_ini = parse_pt_money(valor_str)
            continue
        if lanc_str == "Saldo do dia":
            try:
                saldo_fim = parse_pt_money(valor_str)
            except Exception:
                pass
            continue
        if lanc_str == "S A L D O":
            saldo_fim = parse_pt_money(valor_str)
            continue
        if not data_str:
            continue
        try:
            d = _coerce_date(data_str)
        except Exception:
            continue
        try:
            valor = parse_pt_money(valor_str)
        except Exception:
            continue
        tipo = _normalize_lanc(lanc_str)
        benef = _extract_beneficiario(lanc_str, str(detalhes or ""))
        rows.append({
            "source": "bb",
            "data": d,
            "tipo": tipo,
            "beneficiario": benef,
            "valor": valor,
            "raw_row": {
                "lancamento": lanc_str,
                "detalhes": detalhes,
                "doc": doc,
                "valor_str": valor_str,
                "tipo_lanc": tipo_lanc,
            },
        })
    return pd.DataFrame(rows), {"saldo_inicial": saldo_ini, "saldo_final": saldo_fim}


def _parse_extrato_legacy(ws) -> tuple[pd.DataFrame, dict]:
    """Legacy BB layout: row 3 header; positive values, sign from Historico."""
    header_row = None
    col: dict[str, int] = {}
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if r and r[0] == "Data" and any(c and "istorico" in str(c).lower() for c in r):
            header_row = i
            for j, name in enumerate(r):
                if name:
                    col[str(name).strip().lower()] = j
            break
    if header_row is None:
        raise ValueError("BB legacy: cabeçalho não encontrado.")

    data_c = col.get("data", 0)
    hist_c = col.get("historico", 7)
    valor_c = next(
        (col[k] for k in col if k.startswith("valor")), 8
    )
    num_doc_c = col.get("numero documento", 5)
    cod_hist_c = col.get("cod. historico", 6)

    saldo_ini = saldo_fim = None
    rows = []
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not r or not r[data_c]:
            continue
        hist = str(r[hist_c] or "").strip()
        try:
            d = _coerce_date(r[data_c])
        except Exception:
            continue
        valor_raw = r[valor_c] if valor_c < len(r) else None
        try:
            v = _coerce_money(valor_raw)
        except Exception:
            continue
        if hist == "Saldo Anterior":
            saldo_ini = v
            continue
        if hist == "S A L D O":
            saldo_fim = v
            continue
        sign = _sign_from_historico(hist)
        valor = round(sign * abs(v), 2)
        tipo = _normalize_lanc(hist)
        # Legacy layout has no separate "detalhes"; beneficiario stays blank
        # (will be filled by classifier / manual review).
        rows.append({
            "source": "bb",
            "data": d,
            "tipo": tipo,
            "beneficiario": "",
            "valor": valor,
            "raw_row": {
                "historico": hist,
                "valor_bruto": v,
                "cod_historico": r[cod_hist_c] if cod_hist_c < len(r) else None,
                "num_documento": r[num_doc_c] if num_doc_c < len(r) else None,
            },
        })
    return pd.DataFrame(rows), {"saldo_inicial": saldo_ini, "saldo_final": saldo_fim}


def _detect_layout(ws) -> str:
    """Return 'conta' or 'legacy' by inspecting the first rows."""
    for r in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        if not r:
            continue
        if r[0] == "Data" and len(r) > 1 and str(r[1]).strip() in ("Lançamento", "Lancamento"):
            return "conta"
        if r[0] == "Data" and any(c and "istorico" in str(c).lower() for c in r):
            return "legacy"
    raise ValueError("BB: layout não reconhecido.")


def parse(path) -> tuple[pd.DataFrame, dict]:
    if is_xlsx(path):
        wb = _load_ws(path)
        if "Extrato Conta" in wb.sheetnames:
            return _parse_extrato_conta(wb["Extrato Conta"])
        ws = wb[wb.sheetnames[0]]
    else:
        ws = RowSheet(read_csv_rows(path))
    layout = _detect_layout(ws)
    if layout == "conta":
        return _parse_extrato_conta(ws)
    return _parse_extrato_legacy(ws)
