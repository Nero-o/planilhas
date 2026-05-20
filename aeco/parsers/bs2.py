"""Parser for BS2 extrato — supports both CSV (legacy) and XLSX layouts.

XLSX layout (current banking portal export):
- Single sheet (name varies, e.g. "12. Dezembro - 2025")
- Row 1-2: title + agência/conta
- Row 5: período
- Row 6: ``["Saldo Atual", <valor>]`` — saldo final
- Row 9: header ``Data | Tipo | Detalhe | Banco | Agencia | Conta | Identificador da transação | Identificador da movimentação | Valor | Observação``
- Row 10+: rows. Date is an Excel serial number (or a string with optional BOM);
  one row with ``Tipo == "Saldo"`` and ``Detalhe == "Saldo Inicial"`` marks the
  initial balance. Trailing rows can be variable in length (missing observação).

CSV layout (legacy, kept for retrocompatibility):
- Preamble ``Saldo Atual; R$ X``
- Header line ``Data;Tipo;Detalhe;Banco;Agencia;Conta;Identificador...;Valor;Observação``
- ``Saldo;Saldo Inicial;...`` row for the initial balance
"""
import csv
import io
import re
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
import pandas as pd

from ..normalize import normalize_text, parse_pt_money


_DETAIL_RE = re.compile(r"^[^-]+-\s*\d+\s*-\s*(.+?)\s*$")
_BOM = "﻿"
_EXCEL_EPOCH = datetime(1899, 12, 30)


def _extract_benef(detalhe: str) -> str:
    m = _DETAIL_RE.match(detalhe or "")
    return normalize_text(m.group(1)) if m else normalize_text(detalhe)


def _coerce_date(x):
    """Excel serial number, datetime, or ``DD/MM/YYYY`` string (with optional BOM)."""
    if isinstance(x, datetime):
        return x
    if isinstance(x, (int, float)):
        return _EXCEL_EPOCH + timedelta(days=float(x))
    s = str(x).lstrip(_BOM).strip()
    return datetime.strptime(s, "%d/%m/%Y")


def _coerce_money(x) -> float:
    """Plain number or PT-BR string."""
    if x is None or x == "":
        return 0.0
    if isinstance(x, (int, float)):
        return round(float(x), 2)
    s = str(x).strip()
    try:
        return round(float(s), 2)
    except ValueError:
        return parse_pt_money(s)


def _is_xlsx(path) -> bool:
    if hasattr(path, "name") and isinstance(path.name, str):
        return path.name.lower().endswith(".xlsx")
    if hasattr(path, "read"):
        pos = path.tell() if hasattr(path, "tell") else None
        head = path.read(4)
        if pos is not None and hasattr(path, "seek"):
            path.seek(pos)
        return head[:2] == b"PK"
    return str(path).lower().endswith(".xlsx")


def _parse_xlsx(path) -> tuple[pd.DataFrame, dict]:
    if hasattr(path, "read"):
        if hasattr(path, "seek"):
            path.seek(0)
        wb = openpyxl.load_workbook(io.BytesIO(path.read()), data_only=True)
    else:
        wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    saldo_ini = saldo_fim = None
    header_row = None
    col_idx: dict[str, int] = {}

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row:
            continue
        first = row[0]
        if first == "Saldo Atual" and len(row) > 1 and row[1] is not None:
            try:
                saldo_fim = _coerce_money(row[1])
            except Exception:
                pass
        if first == "Data" and len(row) > 1 and str(row[1]).strip().lower() == "tipo":
            header_row = i
            for j, name in enumerate(row):
                if not name:
                    continue
                key = str(name).strip().lower()
                col_idx[key] = j
            break

    if header_row is None:
        raise ValueError("Cabeçalho do BS2 XLSX não encontrado (esperado 'Data;Tipo;...').")

    data_c = col_idx.get("data", 0)
    tipo_c = col_idx.get("tipo", 1)
    detalhe_c = col_idx.get("detalhe", 2)
    valor_c = col_idx.get("valor", 8)
    obs_c = col_idx.get("observação", col_idx.get("observacao", -1))

    rows = []
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not r or not r[data_c]:
            continue
        tipo_raw = r[tipo_c] if tipo_c < len(r) else None
        detalhe_raw = r[detalhe_c] if detalhe_c < len(r) else None
        valor_raw = r[valor_c] if valor_c < len(r) else None
        if tipo_raw is None:
            continue
        tipo_str = str(tipo_raw).strip()
        if tipo_str == "Saldo":
            try:
                v = _coerce_money(valor_raw)
            except Exception:
                continue
            d = (detalhe_raw or "").strip() if isinstance(detalhe_raw, str) else str(detalhe_raw or "")
            if "Inicial" in d:
                saldo_ini = v
            elif "Final" in d or "Atual" in d:
                saldo_fim = v
            continue
        try:
            data = _coerce_date(r[data_c])
        except Exception:
            continue
        try:
            valor = _coerce_money(valor_raw)
        except Exception:
            continue
        if valor == 0:
            continue
        obs = ""
        if obs_c >= 0 and obs_c < len(r) and r[obs_c] is not None:
            obs = normalize_text(r[obs_c])
        benef = _extract_benef(str(detalhe_raw or ""))
        rows.append({
            "source": "bs2",
            "data": data,
            "tipo": normalize_text(tipo_str).title(),
            "beneficiario": benef,
            "valor": valor,
            "raw_row": {
                "tipo": tipo_str,
                "detalhe": detalhe_raw,
                "valor_str": valor_raw,
                "observacao": obs,
            },
        })
    return pd.DataFrame(rows), {"saldo_inicial": saldo_ini, "saldo_final": saldo_fim}


def _read_lines(path) -> list[str]:
    if hasattr(path, "read"):
        if hasattr(path, "seek"):
            path.seek(0)
        raw = path.read()
        if isinstance(raw, bytes):
            raw = raw.decode("utf-8-sig")
        return [l.lstrip(_BOM) for l in raw.splitlines(keepends=False)]
    with open(path, encoding="utf-8-sig") as f:
        return [l.rstrip("\n").rstrip("\r").lstrip(_BOM) for l in f]


def _parse_csv(path) -> tuple[pd.DataFrame, dict]:
    lines = _read_lines(path)
    saldo_ini = saldo_fim = None
    for l in lines[:15]:
        if l.startswith("Saldo Atual;"):
            saldo_fim = parse_pt_money(l.split(";", 1)[1])
    header_idx = None
    new_format = False
    for i, l in enumerate(lines):
        ls = l.lstrip()
        if ls.startswith("DataContabil") and "Tipo" in l and "Detalhe" in l:
            header_idx, new_format = i, True
            break
        if ls.startswith("Data;") and "Tipo" in l and "Detalhe" in l:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("Cabeçalho do BS2 CSV não encontrado (esperado 'Data;...' ou 'DataContabil;...').")
    rows = []
    reader = csv.reader(lines[header_idx + 1:], delimiter=";")
    for parts in reader:
        if not parts or not parts[0].strip():
            continue
        if "sujeitos a alterações" in parts[0]:
            continue
        if new_format:
            # Normal rows: DataContabil; Data Operação; Tipo; Detalhe; ...(5); Valor; Observação (11 cols)
            # Saldo rows may collapse to 10 cols (single date column).
            if len(parts) >= 3 and parts[1].strip() == "Saldo":
                data_str, tipo, detalhe = parts[0], parts[1], parts[2]
                valor_str = parts[8] if len(parts) > 8 else ""
            elif len(parts) >= 10:
                data_str, tipo, detalhe = parts[0], parts[2], parts[3]
                valor_str = parts[9]
            else:
                continue
        else:
            if len(parts) < 9:
                continue
            data_str, tipo, detalhe = parts[0], parts[1], parts[2]
            valor_str = parts[8]
        if tipo.strip() == "Saldo":
            v = parse_pt_money(valor_str)
            if "Inicial" in detalhe:
                saldo_ini = v
            elif "Final" in detalhe:
                saldo_fim = v
            continue
        d = datetime.strptime(data_str.strip().split(" ")[0], "%d/%m/%Y")
        valor = parse_pt_money(valor_str)
        benef = _extract_benef(detalhe)
        rows.append({
            "source": "bs2",
            "data": d,
            "tipo": normalize_text(tipo).title(),
            "beneficiario": benef,
            "valor": valor,
            "raw_row": {"tipo": tipo, "detalhe": detalhe, "valor_str": valor_str},
        })
    return pd.DataFrame(rows), {"saldo_inicial": saldo_ini, "saldo_final": saldo_fim}


def parse(path) -> tuple[pd.DataFrame, dict]:
    return _parse_xlsx(path) if _is_xlsx(path) else _parse_csv(path)
