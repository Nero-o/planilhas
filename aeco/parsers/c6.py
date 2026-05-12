"""Parser for C6 Bank extrato xlsx.

The bank delivers a password-protected .xlsx (OLE2 EncryptedPackage). The user
must decrypt the file before passing it here (Arquivo → Informações → Proteger
pasta de trabalho → Criptografar com senha → apagar e salvar).

Layout:
- Single sheet named "Transaction".
- Preamble rows (title, agência/conta, "Extrato gerado em ...", period).
- Header row found dynamically by column A == "Data Lançamento":
    A: Data Lançamento  B: Data Contábil  C: Título  D: Descrição
    E: Entrada(R$)  F: Saída(R$)  G: Saldo do Dia(R$)
- No explicit "Saldo Anterior" row; saldo inicial is derived from the first
  row's "Saldo do Dia" minus that day's net movement.
- "Saldo do Dia" repeats for every transaction on the same date (end-of-day
  balance), so the last row's value is the saldo final.
"""
from datetime import datetime
import re

import openpyxl
import pandas as pd

from ..normalize import normalize_text


_PIX_ENVIADO_RE = re.compile(r"^\s*Pix\s+enviado\s+para\s+(.+?)\s*$", re.IGNORECASE)
_PIX_RECEBIDO_RE = re.compile(r"^\s*Pix\s+recebido\s+de\s+(.+?)\s*$", re.IGNORECASE)


def _to_float(x) -> float:
    if x in ("", None):
        return 0.0
    try:
        return float(x)
    except (TypeError, ValueError):
        return 0.0


def _parse_date(x) -> datetime:
    if isinstance(x, datetime):
        return x
    return datetime.strptime(str(x).strip(), "%d/%m/%Y")


def _extract_tipo_benef(titulo: str, descricao: str) -> tuple[str, str]:
    titulo = normalize_text(titulo)
    descricao = normalize_text(descricao)
    m = _PIX_ENVIADO_RE.match(titulo)
    if m:
        return "Pix enviado", normalize_text(m.group(1))
    m = _PIX_RECEBIDO_RE.match(titulo)
    if m:
        return "Pix recebido", normalize_text(m.group(1))
    # Also handle the case where descrição carries the counterparty (e.g.
    # "Pix recebido de X" duplicated). Falls back to titulo as both tipo and
    # beneficiario so the LLM/manual review can resolve.
    return titulo, titulo


def parse(path) -> tuple[pd.DataFrame, dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Transaction"]
    header_row = None
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if r and r[0] == "Data Lançamento":
            header_row = i
            break
    if header_row is None:
        raise ValueError("Could not find C6 header row (expected 'Data Lançamento' in column A)")

    rows = []
    saldo_dia_first = None
    saldo_dia_last = None
    first_date = None
    first_day_net = 0.0
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not r or not r[0]:
            continue
        data_str, _data_contabil, titulo, descricao, entrada, saida, saldo_dia = r[:7]
        d = _parse_date(data_str)
        e = _to_float(entrada)
        s = _to_float(saida)
        valor = round(e - s, 2)
        if valor == 0:
            continue
        tipo, benef = _extract_tipo_benef(titulo, descricao)
        sd = _to_float(saldo_dia)
        if saldo_dia_first is None:
            saldo_dia_first = sd
            first_date = d.date()
        if first_date is not None and d.date() == first_date:
            first_day_net += valor
        saldo_dia_last = sd
        rows.append({
            "source": "c6",
            "data": d,
            "tipo": tipo,
            "beneficiario": benef,
            "valor": valor,
            "raw_row": {
                "titulo": titulo,
                "descricao": descricao,
                "entrada": e,
                "saida": s,
                "saldo_dia": sd,
            },
        })

    saldo_inicial = None
    saldo_final = None
    if saldo_dia_first is not None:
        saldo_inicial = round(saldo_dia_first - first_day_net, 2)
        saldo_final = round(saldo_dia_last, 2)

    return pd.DataFrame(rows), {"saldo_inicial": saldo_inicial, "saldo_final": saldo_final}
