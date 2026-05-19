"""Parser for Conta Simples (cartão corporativo) — xlsx or csv.

Layout (both formats):
- Header row found dynamically (col A = "Data de realização da transação")
- Cols: Data | Status | Data processamento | Tipo | Estabelecimento | Entrada | Saída |
        Cotação | Cartão | Nome do Cartão | Carteira | Token | Categoria | Centro de custo
- "Estabelecimento" has long city/country padding; trim at first 3+ space run.
"""
import re
from datetime import datetime

import openpyxl
import pandas as pd

from ..normalize import normalize_text, parse_pt_money
from ._csv_io import RowSheet, is_xlsx, read_csv_rows


_TRAILING_SPACES_RE = re.compile(r"\s{3,}.*$")


def _clean_estab(s) -> str:
    if s is None:
        return ""
    text = str(s)
    text = _TRAILING_SPACES_RE.sub("", text)
    return normalize_text(text)


def _to_float(x) -> float:
    if x in ("", None):
        return 0.0
    try:
        return float(x)
    except (TypeError, ValueError):
        # PT-BR money string ("1.234,56") — common in CSV exports.
        try:
            return parse_pt_money(str(x))
        except Exception:
            return 0.0


def _coerce_date(x):
    """Datetime pass-through; otherwise parse common BR date strings."""
    if isinstance(x, datetime):
        return x
    s = str(x).strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return x  # leave as-is; downstream will surface the problem


def _load_sheet(path):
    if is_xlsx(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        if "Transações dos cartões" in wb.sheetnames:
            return wb["Transações dos cartões"]
        return wb[wb.sheetnames[0]]
    return RowSheet(read_csv_rows(path))


def parse(path) -> tuple[pd.DataFrame, dict]:
    ws = _load_sheet(path)
    header_row = None
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if r and r[0] == "Data de realização da transação":
            header_row = i
            break
    if header_row is None:
        raise ValueError("Could not find Conta Simples header row")

    rows = []
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not r or not r[0]:
            continue
        data = _coerce_date(r[0])
        status = r[1] if len(r) > 1 else None
        tipo = r[3] if len(r) > 3 else None
        estab = r[4] if len(r) > 4 else None
        entrada = r[5] if len(r) > 5 else None
        saida = r[6] if len(r) > 6 else None
        cartao_nome = r[9] if len(r) > 9 else None
        categoria = r[12] if len(r) > 12 else None
        centro = r[13] if len(r) > 13 else None
        if status and "Recusada" in str(status):
            continue
        e = _to_float(entrada)
        s = _to_float(saida)
        valor = e if e else s
        if valor == 0:
            continue
        rows.append({
            "source": "conta_simples",
            "data": data,
            "tipo": normalize_text(tipo),
            "beneficiario": _clean_estab(estab),
            "valor": round(valor, 2),
            "raw_row": {
                "estabelecimento": estab,
                "cartao": cartao_nome,
                "categoria_origem": categoria,
                "centro_custo": centro,
                "status": status,
            },
        })
    return pd.DataFrame(rows), {}
