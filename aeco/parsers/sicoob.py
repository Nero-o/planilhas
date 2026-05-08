"""Parser for Sicoob extrato xlsx.

Layout:
- Header on row 1: Data | Lançamento | Detalhes | N° documento | Valor | Tipo Lançamento
- Row 2: Saldo Anterior
- Final row: "S A L D O"
- "Saldo do dia" rows interspersed (skipped)
"""
from datetime import datetime
import re
import openpyxl
import pandas as pd

from ..normalize import parse_pt_money, normalize_text


SALDO_TIPOS = {"Saldo Anterior", "Saldo do dia", "S A L D O"}

# Pix: "DD/MM HH:MM [CNPJ-14|CPF-11] NAME"
_PIX_RE = re.compile(
    r"^\s*(?:\d{2}/\d{2}\s+)?(?:\d{2}:\d{2}\s+)?(?:\d{14}\s+|\d{11}\s+)?(.+?)\s*$"
)
# TED: "BANK[3-4 digits] AGENCY[1-4 digits] [CNPJ-14|CPF-11] NAME"
_TED_RE = re.compile(
    r"^\s*\d{3,4}\s+\d{1,4}\s+(?:\d{14}\s+|\d{11}\s+)?(.+?)\s*$"
)


def _extract_beneficiario(lancamento: str, detalhes: str) -> str:
    if not detalhes:
        return ""
    detalhes = detalhes.strip()
    lanc_lower = (lancamento or "").lower()
    if "pix" in lanc_lower:
        m = _PIX_RE.match(detalhes)
        return normalize_text(m.group(1)) if m else normalize_text(detalhes)
    if "ted" in lanc_lower:
        m = _TED_RE.match(detalhes)
        return normalize_text(m.group(1)) if m else normalize_text(detalhes)
    # Pagamento de Boleto / Impostos / Tarifa: use detalhes as-is
    return normalize_text(detalhes)


def _extract_saldos(ws):
    saldo_ini = saldo_fim = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == "Saldo Anterior":
            saldo_ini = parse_pt_money(row[4])
        elif row[1] == "S A L D O":
            saldo_fim = parse_pt_money(row[4])
    return saldo_ini, saldo_fim


def _normalize_tipo(lanc: str) -> str:
    """'Pix - Recebido' -> 'Pix Recebido'; preserves rest."""
    return normalize_text(lanc.replace(" - ", " "))


def parse(path) -> tuple[pd.DataFrame, dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Extrato Conta"]
    saldo_ini, saldo_fim = _extract_saldos(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data_str, lanc, detalhes, doc, valor_str, tipo_lanc = row
        if not lanc or lanc in SALDO_TIPOS:
            continue
        if not data_str:
            continue
        d = datetime.strptime(str(data_str).strip(), "%d/%m/%Y")
        valor = parse_pt_money(valor_str)
        tipo = _normalize_tipo(lanc)
        benef = _extract_beneficiario(lanc, detalhes)
        rows.append({
            "source": "sicoob",
            "data": d,
            "tipo": tipo,
            "beneficiario": benef,
            "valor": valor,
            "raw_row": {
                "lancamento": lanc,
                "detalhes": detalhes,
                "doc": doc,
                "valor_str": valor_str,
                "tipo_lanc": tipo_lanc,
            },
        })
    df = pd.DataFrame(rows)
    return df, {"saldo_inicial": saldo_ini, "saldo_final": saldo_fim}
