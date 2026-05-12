"""Validation: balance reconciliation and duplicate detection vs master."""
from datetime import datetime

import openpyxl
import pandas as pd
from rapidfuzz import fuzz


DETAIL_SHEETS = ["AECO", "SEC", "C6", "TECH", "Conta Simples"]


def _load_master_detail(anual_path: str | None) -> pd.DataFrame:
    if not anual_path:
        return pd.DataFrame(columns=["data", "tipo", "beneficiario", "valor"])
    wb = openpyxl.load_workbook(anual_path, data_only=True)
    rows = []
    for sn in DETAIL_SHEETS:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        header_row = None
        for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
            if r and r[0] == "Data":
                header_row = i
                break
        if not header_row:
            continue
        for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not r or not r[0]:
                continue
            try:
                v = float(r[6]) if r[6] is not None else 0.0
            except (TypeError, ValueError):
                continue
            rows.append({
                "sheet": sn,
                "data": r[0],
                "tipo": str(r[1] or ""),
                "beneficiario": str(r[2] or ""),
                "valor": v,
            })
    return pd.DataFrame(rows)


def _check_saldo(soma_observada: float, expected: float, tol: float = 0.02) -> dict:
    diff = soma_observada - expected
    return {
        "soma_observada": round(soma_observada, 2),
        "esperado": round(expected, 2),
        "diferenca": round(diff, 2),
        "ok": abs(diff) <= tol,
    }


def _detect_unpaired_reversals(
    txs_df: pd.DataFrame, source_name: str, diff: float, tol: float = 0.02
) -> list[dict]:
    """Find transactions whose value matches the saldo discrepancy and lack a
    paired opposite-sign row in the same source.

    Some banks (notably BS2) silently reflect an estorno in the daily saldo
    without emitting the corresponding reversal row. When that happens the
    saldo diff equals the value of the unreverted transaction.

    Pairs entries by absolute value: each opposite-sign row "consumes" one
    candidate. The leftover rows on the side that matches `diff`'s sign are
    flagged as suspects.
    """
    if abs(diff) < tol:
        return []
    src_df = txs_df[txs_df.source == source_name]
    if src_df.empty:
        return []
    abs_target = round(abs(diff), 2)
    matching = src_df[(src_df["valor"].abs() - abs_target).abs() < tol]
    if matching.empty:
        return []
    debits = matching[matching["valor"] < 0]
    credits = matching[matching["valor"] > 0]
    # diff < 0 → missing credit → suspect = leftover debits (was silently estornated)
    # diff > 0 → missing debit  → suspect = leftover credits
    if diff < 0 and len(debits) > len(credits):
        leftover = debits.iloc[len(credits):]
    elif diff > 0 and len(credits) > len(debits):
        leftover = credits.iloc[len(debits):]
    else:
        return []
    return [
        {
            "data": str(r["data"])[:10],
            "tipo": r["tipo"],
            "beneficiario": r["beneficiario"],
            "valor": float(r["valor"]),
            "hint": (
                "Valor coincide com a diferença do saldo. O banco pode ter "
                "estornado este lançamento sem emitir a linha de reversão. "
                "Reexporte o extrato ou adicione o estorno manualmente."
            ),
        }
        for _, r in leftover.iterrows()
    ]


def run(txs_df: pd.DataFrame, saldos: dict, anual_path: str | None = None) -> dict:
    res = {
        "counts": {"total": int(len(txs_df))},
        "saldo": {},
        "saldo_warnings": {},
        "duplicates": [],
    }
    for c in ["green", "yellow", "red"]:
        res["counts"][c] = int((txs_df.confidence == c).sum()) if "confidence" in txs_df else 0

    # Saldos por fonte
    for src, s in saldos.items():
        if not s or s.get("saldo_inicial") is None or s.get("saldo_final") is None:
            continue
        # 'cs' may be the alias for 'conta_simples' from app.py uploads
        source_name = "conta_simples" if src == "cs" else src
        soma = float(txs_df.loc[txs_df.source == source_name, "valor"].sum())
        expected = s["saldo_final"] - s["saldo_inicial"]
        check = _check_saldo(soma, expected)
        res["saldo"][src] = check
        if not check["ok"]:
            suspects = _detect_unpaired_reversals(
                txs_df, source_name, check["diferenca"]
            )
            if suspects:
                res["saldo_warnings"][src] = suspects

    # Duplicates vs master
    master = _load_master_detail(anual_path)
    if not master.empty and not txs_df.empty:
        seen = set()
        # itertuples renames columns starting with `_`; access via dict to be safe
        for _, tx in txs_df.iterrows():
            cands = master[
                (master.data == tx["data"]) & (abs(master.valor - tx["valor"]) < 0.01)
            ]
            for c in cands.itertuples():
                score = fuzz.token_set_ratio(
                    str(tx["beneficiario"]).lower(), str(c.beneficiario).lower()
                )
                if score >= 85:
                    pair = (tx["_id"], c.Index, c.sheet)
                    if pair in seen:
                        continue
                    seen.add(pair)
                    res["duplicates"].append({
                        "new_id": tx["_id"],
                        "data": str(tx["data"])[:10],
                        "valor": float(tx["valor"]),
                        "tipo": tx["tipo"],
                        "beneficiario": tx["beneficiario"],
                        "match_score": int(score),
                        "in_sheet": c.sheet,
                    })
                    break
    return res
