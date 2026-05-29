"""Exporter tests — focus on routing and the C6 'Entrada' header detail."""
import io
from datetime import datetime

import openpyxl
import pandas as pd

from aeco import exporter


def _txs():
    rows = [
        # bb -> SEC tab (empresa is intentionally varied to prove it is ignored)
        {"_id": "1", "source": "bb", "data": datetime(2026, 3, 1),
         "tipo": "Pix Enviado", "beneficiario": "Foo", "valor": -100.0,
         "descricao": "X", "observacoes": "Y", "fluxo_caixa": "F",
         "empresa": "PS", "confidence": "green", "reasoning": "", "classifier": "rule"},
        {"_id": "2", "source": "bb", "data": datetime(2026, 3, 2),
         "tipo": "Pix Enviado", "beneficiario": "Bar", "valor": -200.0,
         "descricao": "X", "observacoes": "Y", "fluxo_caixa": "F",
         "empresa": "Tech", "confidence": "green", "reasoning": "", "classifier": "rule"},
        {"_id": "3", "source": "bb", "data": datetime(2026, 3, 3),
         "tipo": "Pix Recebido", "beneficiario": "Baz", "valor": 1000.0,
         "descricao": "X", "observacoes": "Y", "fluxo_caixa": "F",
         "empresa": "Sec", "confidence": "yellow", "reasoning": "", "classifier": "rule"},
        # bs2 -> TECH tab (empresa=Sec must NOT pull it into SEC)
        {"_id": "4", "source": "bs2", "data": datetime(2026, 3, 4),
         "tipo": "Pix Enviado", "beneficiario": "Dev", "valor": -2000.0,
         "descricao": "X", "observacoes": "Y", "fluxo_caixa": "F",
         "empresa": "Sec", "confidence": "green", "reasoning": "", "classifier": "rule"},
        # conta_simples -> Conta Simples tab
        {"_id": "5", "source": "conta_simples", "data": datetime(2026, 3, 5),
         "tipo": "Compra nacional", "beneficiario": "Microsoft", "valor": -100.0,
         "descricao": "X", "observacoes": "Y", "fluxo_caixa": "F",
         "empresa": "Tech", "confidence": "red", "reasoning": "", "classifier": "rule"},
        # c6 -> C6 tab
        {"_id": "6", "source": "c6", "data": datetime(2026, 3, 6),
         "tipo": "Tarifa", "beneficiario": "Banco", "valor": -50.0,
         "descricao": "X", "observacoes": "Y", "fluxo_caixa": "F",
         "empresa": "Sec", "confidence": "green", "reasoning": "", "classifier": "rule"},
    ]
    return pd.DataFrame(rows)


def _read_xlsx(data: bytes):
    return openpyxl.load_workbook(io.BytesIO(data), data_only=True)


def test_routes_by_source_ignoring_empresa():
    out = exporter.to_xlsx(_txs())
    wb = _read_xlsx(out)
    # conta_simples and c6 route by source
    assert "Conta Simples" in wb.sheetnames
    assert wb["Conta Simples"].cell(row=4, column=3).value == "Microsoft"
    assert wb["C6"].cell(row=4, column=3).value == "Banco"
    # all three bb rows land in SEC, regardless of their Empresa tag
    sec_benefs = [
        wb["SEC"].cell(row=r, column=3).value
        for r in range(4, wb["SEC"].max_row + 1)
    ]
    assert set(sec_benefs) == {"Foo", "Bar", "Baz"}
    # the bs2 row lands in TECH even though its Empresa is "Sec"
    tech_benefs = [
        wb["TECH"].cell(row=r, column=3).value
        for r in range(4, wb["TECH"].max_row + 1)
    ]
    assert tech_benefs == ["Dev"]
    # AECO is the consolidated tab — the union of every transaction
    aeco_benefs = [
        wb["AECO"].cell(row=r, column=3).value
        for r in range(4, wb["AECO"].max_row + 1)
        if wb["AECO"].cell(row=r, column=3).value is not None
    ]
    assert set(aeco_benefs) == {"Foo", "Bar", "Baz", "Dev", "Microsoft", "Banco"}
    assert len(aeco_benefs) == 6  # all rows, none dropped or duplicated


def test_c6_uses_entrada_header():
    out = exporter.to_xlsx(_txs())
    wb = _read_xlsx(out)
    headers = [wb["C6"].cell(row=3, column=c).value for c in range(1, 9)]
    assert headers[6] == "Entrada"  # column 7
    # All other sheets use Valor
    for sheet in ("AECO", "SEC", "TECH", "Conta Simples"):
        assert wb[sheet].cell(row=3, column=7).value == "Valor"


def test_validation_sheet_first():
    out = exporter.to_xlsx(_txs(), saldos={"bb": {"saldo_inicial": 100.0, "saldo_final": -200.0}})
    wb = _read_xlsx(out)
    assert wb.sheetnames[0] == "Validação"
    # Header line
    assert wb["Validação"].cell(row=1, column=1).value == "Fonte"
